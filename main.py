import discord
import json
import os
import datetime
import pytz
import openpyxl
from dotenv import load_dotenv
from discord.ext import commands
from discord.ext import tasks

bot = commands.Bot(command_prefix=">>")
load_dotenv()
with open('Classes.json') as f: 
    classes: dict[str, str] = json.load(f)
ws = openpyxl.load_workbook('Classes.xlsx')
c = ws['Foglio 1']

times = (schedule[0] for schedule in c.values if schedule[0] is not None)

@tasks.loop(time=times)
async def send_link():
    channel = await bot.fetch_channel(821016664610570301)
    today = datetime.datetime.now(tz=pytz.timezone('Europe/Rome')).strftime('%A')
    now = datetime.datetime.now(tz=pytz.timezone('Europe/Rome')).strftime('%H:%M')
    if today not in c.values.__next__():
        return
    else:
        day = c.values.__next__().index(today)
    
    for index, schedule in enumerate(c.values):
        if schedule is None:
            continue
        if schedule == now:
            link = classes[c.cell(index, day)]
            break
    
    await channel.send(link)

@bot.event
async def on_message(message: discord.Message):
    if message.channel.id != 821016664610570301:
        return
    if message.author != bot.user:
        await message.delete()
    await bot.process_commands(message)

@bot.command(name='link')
async def force_send_link(ctx: commands.Context, subject: str):
    subject = subject.title()
    if subject not in classes:
        await ctx.send("Try again")
        return
    
    await ctx.send(classes[subject])

@bot.command(name="orario", aliases=['h'])
async def send_image(ctx: commands.Context):
    await ctx.send(file=discord.File('orario.jpg'))


bot.run(os.getenv('TOKEN'))